"""
报表视图
"""
import logging
import psutil
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from apps.core.response import success_response, error_response
from apps.warehouse.models import Goods, StockIn, StockOut, Warning
from .models import DailyReport

logger = logging.getLogger('apps')


class DashboardView(APIView):
    """仪表盘数据"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        today = datetime.now().date()
        
        # 货物统计
        goods_count = Goods.objects.filter(is_active=True).count()
        warning_goods_count = Goods.objects.filter(
            is_active=True,
            quantity__lte=F('warning_threshold')
        ).count()
        
        # 今日入库统计
        today_in = StockIn.objects.filter(
            stock_in_time__date=today
        ).aggregate(
            count=Count('id'),
            total=Sum('quantity')
        )
        
        # 今日出库统计
        today_out = StockOut.objects.filter(
            stock_out_time__date=today,
            status='completed'
        ).aggregate(
            count=Count('id'),
            total=Sum('quantity')
        )
        
        # 待审批数量
        pending_approval_count = StockOut.objects.filter(status='pending').count()
        
        # 未读预警数量
        unread_warning_count = Warning.objects.filter(is_read=False).count()
        
        # 最近7天入库趋势
        in_trend = []
        out_trend = []
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            in_data = StockIn.objects.filter(
                stock_in_time__date=date
            ).aggregate(total=Sum('quantity'))
            out_data = StockOut.objects.filter(
                stock_out_time__date=date,
                status='completed'
            ).aggregate(total=Sum('quantity'))
            
            in_trend.append({
                'date': date.strftime('%m-%d'),
                'value': float(in_data['total'] or 0)
            })
            out_trend.append({
                'date': date.strftime('%m-%d'),
                'value': float(out_data['total'] or 0)
            })
        
        # 库存预警货物列表
        warning_goods = Goods.objects.filter(
            is_active=True,
            quantity__lte=F('warning_threshold')
        ).values('id', 'name', 'code', 'quantity', 'warning_threshold')[:5]
        
        return success_response(data={
            'goods_count': goods_count,
            'warning_goods_count': warning_goods_count,
            'today_in_count': today_in['count'] or 0,
            'today_in_total': float(today_in['total'] or 0),
            'today_out_count': today_out['count'] or 0,
            'today_out_total': float(today_out['total'] or 0),
            'pending_approval_count': pending_approval_count,
            'unread_warning_count': unread_warning_count,
            'in_trend': in_trend,
            'out_trend': out_trend,
            'warning_goods': list(warning_goods)
        })


class DailyReportView(APIView):
    """每日报表"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).date()
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
        if not end_date:
            end_date = datetime.now().date()
        else:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        reports = []
        current_date = start_date
        
        while current_date <= end_date:
            # 入库统计
            in_data = StockIn.objects.filter(
                stock_in_time__date=current_date
            ).aggregate(
                count=Count('id'),
                total=Sum('quantity')
            )
            
            # 出库统计
            out_data = StockOut.objects.filter(
                stock_out_time__date=current_date,
                status='completed'
            ).aggregate(
                count=Count('id'),
                total=Sum('quantity')
            )
            
            # 预警统计
            warning_count = Warning.objects.filter(
                created_at__date=current_date
            ).count()
            
            reports.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'in_count': in_data['count'] or 0,
                'in_total': float(in_data['total'] or 0),
                'out_count': out_data['count'] or 0,
                'out_total': float(out_data['total'] or 0),
                'warning_count': warning_count
            })
            
            current_date += timedelta(days=1)
        
        return success_response(data=reports)


class ExportView(APIView):
    """数据导出"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        export_type = request.query_params.get('type', 'goods')
        
        wb = Workbook()
        ws = wb.active
        
        # 样式定义
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0066FF', end_color='0066FF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if export_type == 'goods':
            ws.title = '货物清单'
            headers = ['编码', '名称', '品类', '品种', '规格', '库存', '单位', '存放位置', '预警阈值']
            ws.append(headers)
            
            # 设置表头样式
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 数据
            goods_list = Goods.objects.select_related('variety', 'variety__category', 'unit').filter(is_active=True)
            for goods in goods_list:
                ws.append([
                    goods.code,
                    goods.name,
                    goods.variety.category.name if goods.variety else '',
                    goods.variety.name if goods.variety else '',
                    goods.specification,
                    float(goods.quantity),
                    goods.unit.name if goods.unit else '',
                    goods.location,
                    float(goods.warning_threshold)
                ])
            
            filename = f'货物清单_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
            
        elif export_type == 'stock_in':
            ws.title = '入库记录'
            headers = ['货物编码', '货物名称', '入库数量', '单位', '批次号', '供应商', '操作人', '入库时间']
            ws.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            stock_ins = StockIn.objects.select_related('goods', 'goods__unit', 'operator').all()
            for record in stock_ins:
                ws.append([
                    record.goods.code,
                    record.goods.name,
                    float(record.quantity),
                    record.goods.unit.name if record.goods.unit else '',
                    record.batch_no,
                    record.supplier,
                    record.operator.username if record.operator else '',
                    record.stock_in_time.strftime('%Y-%m-%d %H:%M:%S')
                ])
            
            filename = f'入库记录_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
            
        elif export_type == 'stock_out':
            ws.title = '出库记录'
            headers = ['货物编码', '货物名称', '出库数量', '单位', '领用人', '领用部门', '状态', '操作人', '出库时间']
            ws.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            stock_outs = StockOut.objects.select_related('goods', 'goods__unit', 'operator').all()
            for record in stock_outs:
                ws.append([
                    record.goods.code,
                    record.goods.name,
                    float(record.quantity),
                    record.goods.unit.name if record.goods.unit else '',
                    record.receiver,
                    record.receiver_dept,
                    record.get_status_display(),
                    record.operator.username if record.operator else '',
                    record.stock_out_time.strftime('%Y-%m-%d %H:%M:%S') if record.stock_out_time else ''
                ])
            
            filename = f'出库记录_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
        else:
            return error_response(message='无效的导出类型')
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 返回文件
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        logger.info(f"Data exported: {export_type}")
        return response


class SystemMonitorView(APIView):
    """系统资源监控 - 使用psutil"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # CPU使用率
        cpu_percent = psutil.cpu_percent(interval=1)
        cpu_count = psutil.cpu_count()
        
        # 内存使用情况
        memory = psutil.virtual_memory()
        memory_total = round(memory.total / (1024 ** 3), 2)  # GB
        memory_used = round(memory.used / (1024 ** 3), 2)  # GB
        memory_percent = memory.percent
        
        # 磁盘使用情况
        disk = psutil.disk_usage('/')
        disk_total = round(disk.total / (1024 ** 3), 2)  # GB
        disk_used = round(disk.used / (1024 ** 3), 2)  # GB
        disk_percent = disk.percent
        
        # 网络IO
        net_io = psutil.net_io_counters()
        bytes_sent = round(net_io.bytes_sent / (1024 ** 2), 2)  # MB
        bytes_recv = round(net_io.bytes_recv / (1024 ** 2), 2)  # MB
        
        # 系统启动时间
        boot_time = datetime.fromtimestamp(psutil.boot_time())
        uptime = datetime.now() - boot_time
        uptime_str = f"{uptime.days}天{uptime.seconds // 3600}小时"
        
        # 进程数
        process_count = len(psutil.pids())
        
        return success_response(data={
            'cpu': {
                'percent': cpu_percent,
                'count': cpu_count
            },
            'memory': {
                'total': memory_total,
                'used': memory_used,
                'percent': memory_percent
            },
            'disk': {
                'total': disk_total,
                'used': disk_used,
                'percent': disk_percent
            },
            'network': {
                'bytes_sent': bytes_sent,
                'bytes_recv': bytes_recv
            },
            'system': {
                'boot_time': boot_time.strftime('%Y-%m-%d %H:%M:%S'),
                'uptime': uptime_str,
                'process_count': process_count
            }
        })
